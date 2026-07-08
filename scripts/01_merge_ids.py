"""
01_merge_ids.py
---------------
Harmonises sample IDs between the contig-assembly summary and the specimen
metadata, then produces a per-sample result table with:
  - ID          : dot-separated specimen identifier (common denominator)
  - Contig_result: circular | contig (x N) | NA
  - Year        : collection year extracted from metadata

Usage
-----
    python scripts/01_merge_ids.py

Input files (relative to repo root)
-------------------------------------
    data/summary_contigs_mqc.txt   – per-contig assembly output (MitoZ MultiQC)
    data/24032_metadata.xlsx       – specimen metadata from Naturalis collections

Output
------
    output/contig_year_summary.xlsx

ID normalisation logic
----------------------
The metadata SampleID uses dots or colons as separators (e.g. RMNH.INS.1719806,
ZSM:Lep:160416). The assembly summary prefixes each ID with a sequencing-run
code and well position (e.g. NCBN002890_A02_RMNH_INS_1719806), using underscores
throughout.  Matching is done by suffix: the metadata ID is normalised to
underscores and checked against the tail of the assembly ID.

Collection dates stored as raw YYYYMMDD strings (e.g. "18970901") are parsed
to extract the year.
"""

import csv
import datetime
import openpyxl
from collections import defaultdict

# ── paths ────────────────────────────────────────────────────────────────────
CONTIGS_FILE  = "data/summary_contigs_mqc.txt"
METADATA_FILE = "data/24032_metadata.xlsx"
OUTPUT_FILE   = "output/contig_year_summary.xlsx"


# ── helpers ──────────────────────────────────────────────────────────────────
def parse_year(value):
    """Return integer year from a date cell value, or None."""
    if isinstance(value, (datetime.datetime, datetime.date)):
        return value.year
    if isinstance(value, str):
        s = value.strip()
        if len(s) == 8 and s.isdigit():        # YYYYMMDD
            return int(s[:4])
        if len(s) == 4 and s.isdigit():        # YYYY
            return int(s)
    return None


def normalise(specimen_id):
    """Convert dots and colons to underscores for suffix matching."""
    return specimen_id.replace(".", "_").replace(":", "_")


def find_match(summary_id, meta_lookup):
    """Return the metadata key whose normalised form is a suffix of summary_id."""
    for mid, norm in meta_lookup.items():
        if summary_id.endswith(norm):
            return mid
    return None


def dot_form(summary_id):
    """
    Strip the run-code + well prefix and rejoin with dots.
    NCBN002890_A02_RMNH_INS_1719806  →  RMNH.INS.1719806
    """
    parts = summary_id.split("_")
    return ".".join(parts[2:])


# ── load metadata ────────────────────────────────────────────────────────────
wb   = openpyxl.load_workbook(METADATA_FILE)
ws   = wb.active
meta = {}                        # specimen_id → year
for row in ws.iter_rows(min_row=2, values_only=True):
    if not row[0]:
        continue
    meta[row[0]] = parse_year(row[10])   # col 10 = Collection Date

meta_lookup = {mid: normalise(mid) for mid in meta}


# ── load contigs and aggregate per sample ────────────────────────────────────
contig_rows = defaultdict(list)   # summary_id → [label, ...]
with open(CONTIGS_FILE, newline="") as fh:
    for row in csv.DictReader(fh, delimiter="\t"):
        sid    = row["ID"]
        contig = row["Contig"]
        if contig == "NA":
            label = "NA"
        elif "circular" in contig:
            label = "circular"
        else:
            label = "contig"
        contig_rows[sid].append(label)


# ── merge and build result table ─────────────────────────────────────────────
records = []
unmatched = []

for sid, labels in contig_rows.items():
    mid = find_match(sid, meta_lookup)
    if mid is None:
        unmatched.append(sid)
        continue

    year = meta[mid]

    n_circ  = labels.count("circular")
    n_cont  = labels.count("contig")
    n_na    = labels.count("NA")

    if n_na and not n_circ and not n_cont:
        result = "NA"
    else:
        parts = []
        if n_circ:
            parts.append("circular" + (f"x{n_circ}" if n_circ > 1 else ""))
        if n_cont:
            parts.append("contig"   + (f"x{n_cont}" if n_cont  > 1 else ""))
        result = ", ".join(parts)

    records.append((dot_form(sid), result, year if year is not None else ""))

if unmatched:
    print(f"WARNING – {len(unmatched)} IDs could not be matched: {unmatched}")

records.sort(key=lambda r: r[0])

# ── write output ─────────────────────────────────────────────────────────────
out_wb = openpyxl.Workbook()
out_ws = out_wb.active
out_ws.title = "Summary"
out_ws.append(["ID", "Contig_result", "Year"])
for rec in records:
    out_ws.append(list(rec))
for col, width in zip("ABC", [22, 20, 8]):
    out_ws.column_dimensions[col].width = width
out_wb.save(OUTPUT_FILE)

print(f"Written {len(records)} records to {OUTPUT_FILE}")
